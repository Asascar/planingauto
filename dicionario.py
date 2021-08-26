from openpyxl import load_workbook
from tkinter import filedialog
import unicodedata

def lista(linha,dicionario):
    SelecArquivos = filedialog.askopenfilename(title='Dicionário de dados')
    wb = load_workbook(SelecArquivos)
    ws = wb.active
    while linha <= ws.max_row:
        dicionario.append(unicodedata.normalize('NFKC',ws['A' + str(linha)].value).strip())
        linha += 1
    return dicionario    
