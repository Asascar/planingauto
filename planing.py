from openpyxl import load_workbook
from tkinter import filedialog


def planingauto(dicionario):
    SelecArquivos = filedialog.askopenfilenames(title='Planilha a ser filtrada')

    for item in SelecArquivos:
        print(item)
        wb = load_workbook(item)
        ws = wb.active
        col = ws['F']
        linha = 2
    
    def filtro(dicionario,linha):
        while linha <= ws.max_row:
            for item in dicionario:
                try:
                    if item in ws['F' + str(linha)].value:
                        print(ws['F' + str(linha)].value)
                        break
                    else:
                        if item == dicionario[-1]:
                            ws.delete_rows(linha)
                        else:
                            continue
                except:
                    ws.delete_rows(linha)
            linha += 1

    filtro(dicionario,linha)
    filtro(dicionario, linha)
    filtro(dicionario, linha)
    filtro(dicionario, linha)
    filtro(dicionario,linha)
    filtro(dicionario, linha)
    filtro(dicionario, linha)
    filtro(dicionario, linha)            
    wb.save(item)
